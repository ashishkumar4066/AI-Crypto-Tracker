"""
POST /api/v1/import/excel — import Binance Data Download Center Excel exports.

Supports all export types:
  - C2C Order History
  - Deposit History
  - Withdraw History
  - Spot Trade History
  - Spot Order History
  - Convert Order History

Auto-detects the type from row 3. Binance exports have metadata in rows 1-9,
headers in row 10, data from row 11. All files are .xlsx.
"""

from __future__ import annotations

import io
import logging
import re
import uuid
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session

from src.db.database import get_db
from src.db.models import Transaction

logger = logging.getLogger(__name__)
router = APIRouter()


def _uuid() -> str:
    return uuid.uuid4().hex


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _parse_dt(val: Any) -> str:
    if not val:
        return _now_iso()
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.replace(tzinfo=timezone.utc).isoformat()
        except ValueError:
            continue
    return _now_iso()


def _float(val: Any) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


_NUM_UNIT_RE = re.compile(r"^([\d.]+)\s*([A-Za-z]+)$")


def _split_amount_unit(val: Any) -> tuple[float, str]:
    """Parse '82.9USUAL' or '72.123USDT' into (82.9, 'USUAL')."""
    if val is None:
        return 0.0, ""
    s = str(val).strip()
    m = _NUM_UNIT_RE.match(s)
    if m:
        return float(m.group(1)), m.group(2).upper()
    try:
        return float(s), ""
    except ValueError:
        return 0.0, s


def _col_map(headers: list[str]) -> dict[str, int]:
    """Build header_name -> column_index map, skipping empty headers."""
    result = {}
    for i, h in enumerate(headers):
        if h:
            clean = re.sub(r"[^ -~]", "", h).strip()
            if clean in result:
                clean = f"{clean}_{i}"
            result[clean] = i
    return result


def _cell(row: list, col: dict, name: str, default: Any = "") -> Any:
    idx = col.get(name)
    if idx is None or idx >= len(row):
        return default
    return row[idx] if row[idx] is not None else default


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------

def _parse_c2c(headers: list[str], rows: list[list]) -> list[dict]:
    col = _col_map(headers)
    txns = []
    for row in rows:
        status = str(_cell(row, col, "Status")).strip()
        if status != "Completed":
            continue
        order_type = str(_cell(row, col, "Order Type")).strip().lower()
        is_buy = order_type == "buy"
        maker_fee = _float(_cell(row, col, "Maker Fee", 0))
        taker_fee = _float(_cell(row, col, "Taker Fee", 0))
        txns.append({
            "id": _uuid(),
            "datetime": _parse_dt(_cell(row, col, "Created Time")),
            "type": "P2P_BUY" if is_buy else "P2P_SELL",
            "asset": str(_cell(row, col, "Asset")).strip(),
            "amount": _float(_cell(row, col, "Quantity")),
            "fee": maker_fee + taker_fee,
            "fee_asset": str(_cell(row, col, "Fiat Type", "INR")).strip(),
            "price": _float(_cell(row, col, "Price")),
            "quote_amount": _float(_cell(row, col, "Total Price")),
            "counter_asset": str(_cell(row, col, "Fiat Type", "INR")).strip(),
            "source_wallet": "binance_p2p" if is_buy else "binance_spot",
            "dest_wallet": "binance_spot" if is_buy else "binance_p2p",
            "source_endpoint": "excel_c2c",
            "exchange": "binance",
            "external_id": str(_cell(row, col, "Order Number")),
            "status": status,
            "created_at": _now_iso(),
            "updated_at": _now_iso(),
        })
    return txns


def _parse_deposits(headers: list[str], rows: list[list]) -> list[dict]:
    col = _col_map(headers)
    txns = []
    for row in rows:
        txid = str(_cell(row, col, "TXID")).strip()
        if not txid:
            continue
        coin = str(_cell(row, col, "Coin")).strip()
        txns.append({
            "id": _uuid(),
            "datetime": _parse_dt(_cell(row, col, "Time")),
            "type": "DEPOSIT",
            "asset": coin,
            "amount": _float(_cell(row, col, "Amount")),
            "fee": 0.0,
            "fee_asset": coin,
            "network": str(_cell(row, col, "Network")).strip(),
            "address": str(_cell(row, col, "Address")).strip(),
            "txhash": txid,
            "source_wallet": "external",
            "dest_wallet": "binance_spot",
            "source_endpoint": "excel_deposit",
            "exchange": "binance",
            "external_id": txid,
            "status": str(_cell(row, col, "Status")).strip(),
            "created_at": _now_iso(),
            "updated_at": _now_iso(),
        })
    return txns


def _parse_withdrawals(headers: list[str], rows: list[list]) -> list[dict]:
    col = _col_map(headers)
    txns = []
    for row in rows:
        txid = str(_cell(row, col, "TXID")).strip()
        if not txid:
            continue
        coin = str(_cell(row, col, "Coin")).strip()
        addr = str(_cell(row, col, "Address")).strip()
        txns.append({
            "id": _uuid(),
            "datetime": _parse_dt(_cell(row, col, "Time")),
            "type": "WITHDRAWAL",
            "asset": coin,
            "amount": _float(_cell(row, col, "Amount")),
            "fee": _float(_cell(row, col, "Fee")),
            "fee_asset": coin,
            "network": str(_cell(row, col, "Network")).strip(),
            "address": addr,
            "txhash": txid,
            "source_wallet": "binance_spot",
            "dest_wallet": f"external_{addr[:8]}" if addr else "external",
            "source_endpoint": "excel_withdrawal",
            "exchange": "binance",
            "external_id": txid,
            "status": str(_cell(row, col, "Status")).strip(),
            "created_at": _now_iso(),
            "updated_at": _now_iso(),
        })
    return txns


def _parse_spot_trades(headers: list[str], rows: list[list]) -> list[dict]:
    """
    Spot Trade History.
    Headers: Time | Pair | Side | Price | Executed | Amount | Fee
    Values have embedded units: '82.9USUAL', '72.123USDT', '0.0829USUAL'
    """
    col = _col_map(headers)
    txns = []
    for row in rows:
        pair = str(_cell(row, col, "Pair")).strip()
        side = str(_cell(row, col, "Side")).strip().upper()
        price = _float(_cell(row, col, "Price"))
        exec_amt, exec_unit = _split_amount_unit(_cell(row, col, "Executed"))
        total_amt, total_unit = _split_amount_unit(_cell(row, col, "Amount"))
        fee_amt, fee_unit = _split_amount_unit(_cell(row, col, "Fee"))

        if exec_amt == 0:
            continue

        # Base asset = executed unit, quote asset = amount unit
        base_asset = exec_unit or pair.replace(total_unit, "") if total_unit else ""
        quote_asset = total_unit or ""

        txns.append({
            "id": _uuid(),
            "datetime": _parse_dt(_cell(row, col, "Time")),
            "type": "BUY" if side == "BUY" else "SELL",
            "asset": base_asset,
            "amount": exec_amt,
            "fee": fee_amt,
            "fee_asset": fee_unit,
            "price": price,
            "quote_amount": total_amt,
            "counter_asset": quote_asset,
            "source_wallet": "binance_spot",
            "dest_wallet": "binance_spot",
            "source_endpoint": "excel_spot_trade",
            "exchange": "binance",
            "external_id": f"{pair}_{_cell(row, col, 'Time')}_{exec_amt}",
            "status": "COMPLETED",
            "created_at": _now_iso(),
            "updated_at": _now_iso(),
        })
    return txns


def _parse_spot_orders(headers: list[str], rows: list[list]) -> list[dict]:
    """
    Spot Order History.
    Headers: Time | OrderNo | Pair | Type | Side | Order Price | Order Amount |
             Time | Executed | Average Price | Trading total | Status
    """
    col = _col_map(headers)
    txns = []
    for row in rows:
        status = str(_cell(row, col, "Status")).strip().upper()
        if status not in ("FILLED", "PARTIALLY_FILLED"):
            continue

        pair = str(_cell(row, col, "Pair")).strip()
        side = str(_cell(row, col, "Side")).strip().upper()
        avg_price = _float(_cell(row, col, "Average Price"))

        # Headers may have invisible BOM chars — find fuzzy match
        exec_key = next((k for k in col if k.startswith("Executed")), "Executed")
        total_key = next((k for k in col if k.startswith("Trading total")), "Trading total")

        exec_amt, exec_unit = _split_amount_unit(_cell(row, col, exec_key))
        total_amt, total_unit = _split_amount_unit(_cell(row, col, total_key))
        order_no = str(_cell(row, col, "OrderNo")).strip()

        if exec_amt == 0:
            continue

        base_asset = exec_unit or ""
        quote_asset = total_unit or ""

        txns.append({
            "id": _uuid(),
            "datetime": _parse_dt(_cell(row, col, "Time")),
            "type": "BUY" if side == "BUY" else "SELL",
            "asset": base_asset,
            "amount": exec_amt,
            "fee": 0.0,
            "fee_asset": "",
            "price": avg_price,
            "quote_amount": total_amt,
            "counter_asset": quote_asset,
            "source_wallet": "binance_spot",
            "dest_wallet": "binance_spot",
            "source_endpoint": "excel_spot_order",
            "exchange": "binance",
            "external_id": order_no,
            "order_id": order_no,
            "status": status,
            "created_at": _now_iso(),
            "updated_at": _now_iso(),
        })
    return txns


def _parse_convert(headers: list[str], rows: list[list]) -> list[dict]:
    """
    Convert Order History.
    Headers: Time | OrderNo | Pair | Side | Price | Executed | Amount | Fee | Status
    (structure similar to spot trades)
    """
    col = _col_map(headers)
    txns = []
    for row in rows:
        status = str(_cell(row, col, "Status")).strip()
        if status not in ("Filled", "Completed", "SUCCESS", "Success"):
            continue

        exec_amt, exec_unit = _split_amount_unit(_cell(row, col, "Executed"))
        total_amt, total_unit = _split_amount_unit(_cell(row, col, "Amount"))
        order_no = str(_cell(row, col, "OrderNo", _cell(row, col, "Order No"))).strip()

        if exec_amt == 0:
            continue

        txns.append({
            "id": _uuid(),
            "datetime": _parse_dt(_cell(row, col, "Time")),
            "type": "CONVERT",
            "asset": exec_unit,
            "amount": exec_amt,
            "fee": 0.0,
            "fee_asset": "",
            "price": _float(_cell(row, col, "Price")),
            "quote_amount": total_amt,
            "counter_asset": total_unit,
            "source_wallet": "binance_convert",
            "dest_wallet": "binance_spot",
            "source_endpoint": "excel_convert",
            "exchange": "binance",
            "external_id": order_no or f"convert_{_cell(row, col, 'Time')}_{exec_amt}",
            "status": status,
            "created_at": _now_iso(),
            "updated_at": _now_iso(),
        })
    return txns


def _parse_transaction_history(headers: list[str], rows: list[list]) -> list[dict]:
    """
    Transaction History — the master ledger export from Binance.
    Headers: User ID | Time | Account | Operation | Coin | Change | Remark

    Operation types map to transaction types:
      Transaction Buy / Transaction Spend / Transaction Sold / Transaction Fee
      P2P Trading, Deposit, Withdraw, Binance Convert,
      Small Assets Exchange BNB, Distribution, Airdrop Assets,
      Transfer Between Spot and Funding, etc.
    """
    col = _col_map(headers)
    txns = []

    _OP_TYPE_MAP = {
        "Transaction Buy": "BUY",
        "Transaction Spend": "SELL",
        "Transaction Sold": "SELL",
        "Transaction Fee": "FEE",
        "Transaction Revenue": "REWARD",
        "P2P Trading": "P2P_BUY",
        "Deposit": "DEPOSIT",
        "Withdraw": "WITHDRAWAL",
        "Binance Convert": "CONVERT",
        "Small Assets Exchange BNB": "DUST_CONVERSION",
        "Distribution": "REWARD",
        "Airdrop Assets": "REWARD",
        "Launchpool Airdrop - User Claim Distribution": "REWARD",
        "Launchpool Subscription/Redemption": "STAKING",
        "Commission Rebate": "REWARD",
        "Crypto Box": "REWARD",
        "Token Swap - Distribution": "CONVERT",
        "Token Swap - Redenomination/Rebranding": "CONVERT",
        "Simple Earn Flexible Subscription": "STAKING",
        "Simple Earn Flexible Redemption": "STAKING",
        "Transfer Between Spot and Funding": "INTERNAL_TRANSFER",
        "Transfer Between Spot and UM Futures": "INTERNAL_TRANSFER",
        "Asset Recovery": "REWARD",
        "Fee": "FEE",
        "Funding Fee": "FEE",
        "Realized Profit and Loss": "REWARD",
    }

    for row in rows:
        operation = str(_cell(row, col, "Operation")).strip()
        if not operation:
            continue

        coin = str(_cell(row, col, "Coin")).strip()
        change_val = _float(_cell(row, col, "Change"))
        account = str(_cell(row, col, "Account")).strip()
        remark = str(_cell(row, col, "Remark")).strip()
        time_str = str(_cell(row, col, "Time")).strip()

        txn_type = _OP_TYPE_MAP.get(operation, "UNKNOWN")

        # P2P Trading: positive change = buy, negative = sell
        if operation == "P2P Trading":
            txn_type = "P2P_BUY" if change_val >= 0 else "P2P_SELL"

        # Determine wallet based on account
        wallet = "binance_spot"
        if account == "Funding":
            wallet = "binance_funding"
        elif account == "UM Futures":
            wallet = "binance_futures"

        src_wallet = wallet
        dst_wallet = wallet
        if txn_type == "DEPOSIT":
            src_wallet = "external"
            dst_wallet = wallet
        elif txn_type == "WITHDRAWAL":
            src_wallet = wallet
            dst_wallet = "external"
        elif txn_type == "INTERNAL_TRANSFER":
            if change_val > 0:
                src_wallet = "binance_funding"
                dst_wallet = "binance_spot"
            else:
                src_wallet = "binance_spot"
                dst_wallet = "binance_funding"

        # Build a unique external_id from time + operation + coin + change
        ext_id = f"txnhist_{time_str}_{operation}_{coin}_{change_val}"

        txns.append({
            "id": _uuid(),
            "datetime": _parse_dt(time_str),
            "type": txn_type,
            "asset": coin,
            "amount": abs(change_val),
            "fee": 0.0,
            "fee_asset": coin if txn_type == "FEE" else "",
            "price": 0.0,
            "quote_amount": 0.0,
            "counter_asset": "",
            "source_wallet": src_wallet,
            "dest_wallet": dst_wallet,
            "source_endpoint": "excel_transaction_history",
            "exchange": "binance",
            "external_id": ext_id,
            "status": "COMPLETED",
            "raw_json": remark if remark else None,
            "created_at": _now_iso(),
            "updated_at": _now_iso(),
        })

    return txns


# ---------------------------------------------------------------------------
# Type detection & parser routing
# ---------------------------------------------------------------------------

_PARSERS = {
    "C2C": _parse_c2c,
    "Deposit": _parse_deposits,
    "Withdraw": _parse_withdrawals,
    "Spot Trade": _parse_spot_trades,
    "Spot Order": _parse_spot_orders,
    "Convert": _parse_convert,
    "Transaction": _parse_transaction_history,
}


def _detect_type(row3_text: str) -> str | None:
    """Match row 3 label to a parser key."""
    for key in _PARSERS:
        if key.lower().replace(" ", "") in row3_text.lower().replace(" ", ""):
            return key
    return None


# ---------------------------------------------------------------------------
# Route
# ---------------------------------------------------------------------------

def _process_one_file(file_content: bytes, filename: str, db: Session) -> dict:
    """Parse and import a single Excel file. Returns a result dict."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_content))
    ws = wb.active

    all_rows = []
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True
    ):
        all_rows.append(list(row))
    wb.close()

    if len(all_rows) < 11:
        return {"file": filename, "status": "error", "error": "File too short"}

    row3_text = " ".join(str(v).strip() for v in all_rows[2] if v is not None)
    detected = _detect_type(row3_text)

    if not detected:
        return {"file": filename, "status": "error", "error": f"Unknown type: '{row3_text}'"}

    headers = [str(v or "").strip() for v in all_rows[9]]

    data_rows = []
    for row in all_rows[10:]:
        values = list(row)
        non_empty = [v for v in values if v is not None and str(v).strip()]
        if not non_empty:
            continue
        row_text = " ".join(str(v or "") for v in values)
        if "No data" in row_text:
            continue
        data_rows.append(values)

    parser = _PARSERS[detected]
    txns = parser(headers, data_rows)

    inserted = 0
    skipped = 0
    for t in txns:
        existing = (
            db.query(Transaction.id)
            .filter_by(
                exchange=t.get("exchange"),
                source_endpoint=t.get("source_endpoint"),
                external_id=t.get("external_id"),
            )
            .first()
        )
        if existing:
            skipped += 1
            continue

        row = Transaction(
            id=t["id"],
            datetime=t.get("datetime"),
            type=t.get("type"),
            asset=t.get("asset"),
            amount=t.get("amount", 0),
            fee=t.get("fee", 0),
            fee_asset=t.get("fee_asset"),
            price=t.get("price"),
            quote_amount=t.get("quote_amount"),
            counter_asset=t.get("counter_asset"),
            txhash=t.get("txhash"),
            network=t.get("network"),
            address=t.get("address"),
            source_wallet=t.get("source_wallet"),
            dest_wallet=t.get("dest_wallet"),
            source_endpoint=t.get("source_endpoint"),
            exchange=t.get("exchange", "binance"),
            external_id=t.get("external_id"),
            order_id=t.get("order_id"),
            status=t.get("status"),
            needs_review=False,
            created_at=t.get("created_at"),
            updated_at=t.get("updated_at"),
        )
        db.add(row)
        inserted += 1

    if inserted > 0:
        db.commit()

    logger.info(
        "Import: file=%s, type=%s, parsed=%d, inserted=%d, skipped=%d",
        filename, detected, len(txns), inserted, skipped,
    )

    return {
        "file": filename,
        "status": "success",
        "type_detected": detected,
        "rows_in_file": len(data_rows),
        "rows_parsed": len(txns),
        "inserted": inserted,
        "skipped": skipped,
    }


@router.post("/import/excel")
async def import_excel(
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
):
    """Import one or more Binance Data Download Center Excel files."""
    results = []
    total_inserted = 0
    total_skipped = 0
    total_parsed = 0

    for f in files:
        if not f.filename or not f.filename.endswith((".xlsx", ".xls")):
            results.append({"file": f.filename, "status": "error", "error": "Not an .xlsx file"})
            continue

        content = await f.read()
        try:
            result = _process_one_file(content, f.filename, db)
        except Exception as exc:
            logger.exception("Failed to process %s", f.filename)
            result = {"file": f.filename, "status": "error", "error": str(exc)}

        results.append(result)
        if result.get("status") == "success":
            total_inserted += result.get("inserted", 0)
            total_skipped += result.get("skipped", 0)
            total_parsed += result.get("rows_parsed", 0)

    return {
        "files_processed": len(results),
        "total_parsed": total_parsed,
        "total_inserted": total_inserted,
        "total_skipped": total_skipped,
        "results": results,
    }


@router.get("/import/tally")
def import_tally(
    fy: str | None = None,
    asset: str | None = None,
    db: Session = Depends(get_db),
):
    """
    Compare API-synced vs Excel-imported data with breakdowns.
    Supports filtering by FY and/or asset.
    """
    from sqlalchemy import func
    from src.config import FY_DATE_RANGES

    q = db.query(Transaction)

    if fy and fy in FY_DATE_RANGES:
        start_dt, end_dt = FY_DATE_RANGES[fy]
        q = q.filter(
            Transaction.datetime >= start_dt.isoformat(),
            Transaction.datetime <= end_dt.isoformat(),
        )

    if asset:
        q = q.filter(Transaction.asset == asset.upper())

    # --- by source_endpoint + type ---
    source_rows = (
        q.with_entities(
            Transaction.source_endpoint,
            Transaction.type,
            func.count(Transaction.id),
            func.sum(Transaction.amount),
        )
        .group_by(Transaction.source_endpoint, Transaction.type)
        .all()
    )

    api_sources = {}
    excel_sources = {}
    for ep, txn_type, cnt, total_amt in source_rows:
        is_excel = ep.startswith("excel_")
        bucket = excel_sources if is_excel else api_sources
        if ep not in bucket:
            bucket[ep] = {"total": 0, "types": {}}
        bucket[ep]["total"] += cnt
        bucket[ep]["types"][txn_type] = {
            "count": cnt,
            "amount": round(total_amt or 0, 8),
        }

    # --- by asset (top coins) ---
    asset_rows = (
        q.with_entities(
            Transaction.asset,
            Transaction.source_endpoint,
            func.count(Transaction.id),
        )
        .group_by(Transaction.asset, Transaction.source_endpoint)
        .all()
    )

    coins = {}
    for coin, ep, cnt in asset_rows:
        if coin not in coins:
            coins[coin] = {"api": 0, "excel": 0}
        if ep.startswith("excel_"):
            coins[coin]["excel"] += cnt
        else:
            coins[coin]["api"] += cnt

    # --- by type (aggregated) ---
    type_rows = (
        q.with_entities(
            Transaction.type,
            Transaction.source_endpoint,
            func.count(Transaction.id),
        )
        .group_by(Transaction.type, Transaction.source_endpoint)
        .all()
    )

    by_type = {}
    for txn_type, ep, cnt in type_rows:
        if txn_type not in by_type:
            by_type[txn_type] = {"api": 0, "excel": 0}
        if ep.startswith("excel_"):
            by_type[txn_type]["excel"] += cnt
        else:
            by_type[txn_type]["api"] += cnt

    # --- available filter values ---
    all_assets = sorted(set(
        r[0] for r in db.query(Transaction.asset).distinct().all() if r[0]
    ))

    return {
        "filters": {"fy": fy, "asset": asset},
        "totals": {
            "api": sum(v["total"] for v in api_sources.values()),
            "excel": sum(v["total"] for v in excel_sources.values()),
        },
        "by_source": {"api": api_sources, "excel": excel_sources},
        "by_type": by_type,
        "by_asset": coins,
        "available_assets": all_assets,
    }
